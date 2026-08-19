"""Import the 申報規範 spreadsheet finance already maintains.

Imported cells count as human-authored, so extraction will not overwrite them:
a person wrote this, and a later model run does not get to silently disagree.
They carry no citations, so they are flagged for review until someone anchors
them to provisions — the matrix is only self-maintaining for cells that say
where they came from.

Column headers are matched loosely, because the sheet in circulation has
headers spanning several lines and mixing 繁體/簡體.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from taxwatch.models import (
    FieldSource,
    RequirementField,
    RequirementStatus,
    TaxRequirement,
)
from taxwatch.requirements.fields import FIELD_KEYS
from taxwatch.taxonomy import UNCLASSIFIED, classify

logger = logging.getLogger(__name__)


class MissingDependency(RuntimeError):
    """Raised when the optional spreadsheet reader is not installed."""


# Header text → field key. Matched as substrings against a normalised header,
# so 「應納稅額計算公式 (財務簡式)」 still lands on `formula`.
_HEADER_HINTS: tuple[tuple[str, str], ...] = (
    # Specific / multi-word hints first
    ("tax type", "_tax"),
    ("稅種", "_tax"),
    ("税种", "_tax"),
    ("tax scenario", "_scenario"),
    ("sub-item", "_scenario"),
    ("subitem", "_scenario"),
    ("課稅情境", "_scenario"),
    ("子項目", "_scenario"),
    ("子项目", "_scenario"),
    ("withholding agent", "_role"),
    ("taxpayer", "_role"),
    ("角色", "_role"),
    ("requirement", "applicability"),
    ("適用條件", "applicability"),
    ("tax event", "taxable_event"),
    ("trigger point", "taxable_event"),
    ("課稅事件", "taxable_event"),
    ("触发时点", "taxable_event"),
    ("觸發時點", "taxable_event"),
    ("statutory rate", "rate"),
    ("rate", "rate"),
    ("稅率", "rate"),
    ("税率", "rate"),
    ("taxable item", "taxable_items"),
    ("應稅項目", "taxable_items"),
    ("应税项目", "taxable_items"),
    ("calculation formula", "formula"),
    ("formula", "formula"),
    ("計算公式", "formula"),
    ("计算公式", "formula"),
    ("tax base", "tax_base"),
    ("稅基", "tax_base"),
    ("税基", "tax_base"),
    ("deduction", "deductions"),
    ("credit", "deductions"),
    ("扣除", "deductions"),
    ("扣抵", "deductions"),
    ("incentive", "incentives"),
    ("reduction", "incentives"),
    ("租稅優惠", "incentives"),
    ("租税优惠", "incentives"),
    ("filing deadline", "filing_deadline"),
    ("申報期限", "filing_deadline"),
    ("申报期限", "filing_deadline"),
    ("payment deadline", "payment_deadline"),
    ("collection period", "payment_deadline"),
    ("繳款期限", "payment_deadline"),
    ("缴款期限", "payment_deadline"),
    ("collection management", "administration"),
    ("徵收管理", "administration"),
    ("征收管理", "administration"),
    ("policy basis", "_policy_basis"),
    ("change content", "_change_content"),
)


def import_workbook(
    session: Session,
    path: str | Path,
    *,
    country: str = "CN",
    sheet: str | int = 0,
    source_note: str = "",
) -> dict[str, Any]:
    """Load a 申報規範 sheet into the database.

    Rows are keyed on (tax type, scenario, role) like everything else, so
    re-importing a corrected sheet updates in place rather than duplicating.
    """
    rows = _read_rows(path, sheet)
    if not rows:
        return {
            "rows": 0,
            "imported": 0,
            "skipped": 0,
            "columns_mapped": [],
            "unmapped_headers": [],
        }

    header, *body = rows
    mapping = _map_columns(header)
    if "_scenario" not in mapping.values():
        raise ValueError("No 子項目/課稅情境 column found — cannot key the rows")

    unmapped_headers = [
        header[i].strip() for i in range(len(header)) if i not in mapping and header[i].strip()
    ]

    imported = 0
    skipped = 0

    for raw in body:
        record = _row_to_record(raw, mapping)
        if not record.get("_scenario"):
            skipped += 1
            continue
        _upsert(session, record, country=country, source_note=source_note)
        imported += 1

    session.commit()
    logger.info("Imported %d requirement rows from %s", imported, path)
    return {
        "rows": len(body),
        "imported": imported,
        "skipped": skipped,
        "columns_mapped": sorted({v for v in mapping.values() if not v.startswith("_")}),
        "unmapped_headers": unmapped_headers,
    }


# ---------- internals ----------


def _read_rows(path: str | Path, sheet: str | int = 0) -> list[list[str]]:
    p = Path(path)
    ext = p.suffix.lower()

    if ext in (".md", ".markdown"):
        return _read_markdown_table(p)
    elif ext == ".csv":
        return _read_csv(p)
    elif ext == ".xlsx":
        return _read_xlsx(p, sheet)
    else:
        # Fallback to xlsx if extension not recognized or try guessing
        return _read_xlsx(p, sheet)


def _read_xlsx(path: Path, sheet: str | int) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional extra
        raise MissingDependency("Reading .xlsx needs openpyxl: pip install -e '.[xlsx]'") from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    worksheet = workbook[sheet] if isinstance(sheet, str) else workbook.worksheets[sheet]
    rows = [
        ["" if cell is None else str(cell).strip() for cell in row]
        for row in worksheet.iter_rows(values_only=True)
    ]
    workbook.close()
    return [row for row in rows if any(cell for cell in row)]


def _read_csv(path: Path) -> list[list[str]]:
    import csv

    with open(path, mode="r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        return [[cell.strip() for cell in row] for row in reader if any(c.strip() for c in row)]


def _read_markdown_table(path: Path) -> list[list[str]]:
    with open(path, mode="r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    rows: list[list[str]] = []
    _BR_RE = re.compile(r"<\s*br\s*/?\s*>", re.IGNORECASE)

    for line in content.splitlines():
        line_clean = line.strip()
        if not line_clean or not line_clean.startswith("|") or not line_clean.endswith("|"):
            continue

        # Strip outer pipes
        inner = line_clean[1:-1]
        raw_cells = inner.split("|")

        cleaned_cells: list[str] = []
        is_delimiter_row = True

        for c in raw_cells:
            cell_text = c.strip()
            # Check if this cell is part of markdown delimiter like :---: or ---
            stripped_dashes = cell_text.replace("-", "").replace(":", "").strip()
            if stripped_dashes:
                is_delimiter_row = False

            # Transform <br> to newline
            cell_text = _BR_RE.sub("\n", cell_text)
            # Remove markdown bold **
            cell_text = cell_text.replace("**", "")
            cleaned_cells.append(cell_text.strip())

        if is_delimiter_row:
            continue

        if any(cleaned_cells):
            rows.append(cleaned_cells)

    return rows


def _map_columns(header: list[str]) -> dict[int, str]:
    """Header index → field key, best-effort."""
    mapping: dict[int, str] = {}
    for index, cell in enumerate(header):
        raw_cell = cell.strip().lower()
        no_space = re.sub(r"[\s\-_]+", "", raw_cell)
        for hint, field_key in _HEADER_HINTS:
            hint_lower = hint.lower()
            hint_no_space = re.sub(r"[\s\-_]+", "", hint_lower)
            if hint_lower in raw_cell or hint_no_space in no_space:
                mapping[index] = field_key
                break
    return mapping


def _row_to_record(row: list[str], mapping: dict[int, str]) -> dict[str, str]:
    record: dict[str, str] = {}
    for index, field_key in mapping.items():
        if index < len(row):
            record[field_key] = row[index].strip()
    return record


def _upsert(
    session: Session,
    record: dict[str, str],
    *,
    country: str,
    source_note: str = "",
) -> TaxRequirement:
    tax_key = _tax_key(record.get("_tax", ""), country=country)
    scenario = record["_scenario"]
    role = record.get("_role", "")

    requirement = (
        session.query(TaxRequirement)
        .filter_by(country=country, tax_key=tax_key, scenario=scenario, taxpayer_role=role)
        .first()
    )
    if requirement is None:
        requirement = TaxRequirement(
            country=country,
            tax_key=tax_key,
            scenario=scenario,
            taxpayer_role=role,
        )
        session.add(requirement)
    requirement.status = RequirementStatus.DRAFT

    if source_note.strip():
        note_str = source_note.strip()
        if requirement.notes:
            if note_str not in requirement.notes:
                requirement.notes = f"{requirement.notes}\n{note_str}"
        else:
            requirement.notes = note_str

    session.flush()

    existing = {f.field_key: f for f in requirement.fields}
    for field_key in FIELD_KEYS:
        value = record.get(field_key, "").strip()
        if not value:
            continue

        field = existing.get(field_key)
        if field is None:
            field = RequirementField(requirement_id=requirement.id, field_key=field_key)
            session.add(field)

        field.value = value
        field.source = FieldSource.IMPORT
        field.citations = []
        field.confidence = 0.0
        # Imported prose has no provision behind it yet, so the system cannot
        # tell when it goes out of date. Say so rather than imply it is tracked.
        field.needs_review = True
        field.review_reason = "由試算表匯入，尚未對應條文，法規異動時無法自動追蹤"

    session.flush()
    return requirement


def _tax_key(label: str, country: str = "CN") -> str:
    return classify(label, country=country).key
